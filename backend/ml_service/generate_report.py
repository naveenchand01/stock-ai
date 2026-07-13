"""
===========================================================================
  DISSERTATION REPORT GENERATOR
  Generates charts, tables and a full HTML report from NIFTY50 master CSV
===========================================================================
  Outputs (all saved to dissertation_results/report/):
    figures/
      01_model_comparison_mape.png       - MAPE bar chart per model
      02_model_comparison_rmse_r2.png    - RMSE & R² grouped bar chart
      03_r2_heatmap.png                  - R² heatmap: stocks × models
      04_mape_heatmap.png                - MAPE heatmap: stocks × models
      05_directional_accuracy.png        - Directional accuracy bar chart
      06_f1_score_heatmap.png            - F1-Score heatmap
      07_best_model_r2_top15.png         - Top 15 stocks by R² (CNN-LSTM)
      08_cnnlstm_scatter_r2.png          - R² scatter: CNN-LSTM vs LSTM
    dissertation_table.xlsx              - Excel workbook with formatted tables
    dissertation_report.html             - Full HTML report (open in browser)

  Usage:
    cd backend/ml_service
    python generate_report.py
===========================================================================
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")          # Non-interactive backend (no display needed)
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.patches as mpatches

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
MASTER_CSV  = os.path.join("dissertation_results", "NIFTY50_master_dissertation_metrics.csv")
REPORT_DIR  = os.path.join("dissertation_results", "report")
FIG_DIR     = os.path.join(REPORT_DIR, "figures")
EXCEL_PATH  = os.path.join(REPORT_DIR, "dissertation_table.xlsx")
HTML_PATH   = os.path.join(REPORT_DIR, "dissertation_report.html")

os.makedirs(FIG_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Design constants
# ---------------------------------------------------------------------------
MODEL_ORDER  = ["ARIMA", "SARIMA", "ARIMA-LSTM", "XGBoost", "LSTM", "CNN-LSTM"]
MODEL_COLORS = {
    "ARIMA":         "#A8C4E0",
    "SARIMA":        "#6FA8D0",
    "ARIMA-LSTM":    "#F4A460",
    "XGBoost":       "#E07840",
    "LSTM":          "#8CC0A0",
    "CNN-LSTM":      "#2E8B57",
}
FONT_TITLE  = {"fontsize": 13, "fontweight": "bold", "color": "#1a1a2e"}
FONT_AXIS   = {"fontsize": 10, "color": "#333333"}
GRID_COLOR  = "#e8e8e8"
BG_COLOR    = "#fafafa"

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         True,
    "grid.color":        GRID_COLOR,
    "grid.linewidth":    0.6,
    "figure.facecolor":  BG_COLOR,
    "axes.facecolor":    BG_COLOR,
})

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
print("Loading master CSV …")
df_raw = pd.read_csv(MASTER_CSV)

# Remove the phantom "RELIANCE_dissertation_metrics" row if present
df_raw = df_raw[~df_raw["Symbol"].str.contains("dissertation_metrics", na=False)]
df_raw = df_raw[~df_raw["Symbol"].isin(["TMCV"])]   # too little data, unreliable

# Keep only proper Nifty 50 models
df_raw = df_raw[df_raw["Model"].isin(MODEL_ORDER)]

# Calculate Forecast Accuracy = 100 - MAPE (clipped at 0)
df_raw["Accuracy (%)"] = (100 - df_raw["MAPE"]).clip(lower=0)

print(f"  Symbols : {df_raw['Symbol'].nunique()}")
print(f"  Models  : {sorted(df_raw['Model'].unique())}")

# Pivot helper
def pivot(metric):
    return df_raw.pivot_table(index="Symbol", columns="Model", values=metric, aggfunc="first")

pv_rmse  = pivot("RMSE")
pv_mape  = pivot("MAPE")
pv_r2    = pivot("R2")
pv_acc   = pivot("Accuracy (%)")
pv_prec  = pivot("Precision")
pv_rec   = pivot("Recall")
pv_f1    = pivot("F1_Score")

# Ensure column order
for pv in [pv_rmse, pv_mape, pv_r2, pv_acc, pv_prec, pv_rec, pv_f1]:
    for col in MODEL_ORDER:
        if col not in pv.columns:
            pv[col] = np.nan
    pv = pv[MODEL_ORDER]

# CNN-LSTM best-model frame
cnn = df_raw[df_raw["Model"] == "CNN-LSTM"].set_index("Symbol")

# ---------------------------------------------------------------------------
# Figure 1 – Average MAPE by Model (bar chart)
# ---------------------------------------------------------------------------
print("\n[1/8] Average MAPE by model …")
fig, ax = plt.subplots(figsize=(9, 5))
avg_mape = df_raw.groupby("Model")["MAPE"].mean().reindex(MODEL_ORDER)
bars = ax.barh(MODEL_ORDER,
               avg_mape.values,
               color=[MODEL_COLORS[m] for m in MODEL_ORDER],
               edgecolor="white", linewidth=0.8)
for bar, val in zip(bars, avg_mape.values):
    ax.text(val + 0.1, bar.get_y() + bar.get_height()/2,
            f"{val:.2f}%", va="center", fontsize=9, color="#333")
ax.set_xlabel("Average MAPE (%)", **FONT_AXIS)
ax.set_title("Average MAPE by Model — Nifty 50 (80/20 split)", **FONT_TITLE)
ax.invert_yaxis()
plt.tight_layout()
p1 = os.path.join(FIG_DIR, "01_model_comparison_mape.png")
plt.savefig(p1, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p1}")

# ---------------------------------------------------------------------------
# Figure 2 – Average RMSE & R² grouped bar chart
# ---------------------------------------------------------------------------
print("[2/8] RMSE & R² grouped …")
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
avg_rmse = df_raw.groupby("Model")["RMSE"].mean().reindex(MODEL_ORDER)
avg_r2   = df_raw.groupby("Model")["R2"].mean().reindex(MODEL_ORDER)

for ax, metric, vals, label, fmt in [
    (axes[0], "Avg RMSE (₹)", avg_rmse, "Lower is better", "{:.1f}"),
    (axes[1], "Avg R² Score",  avg_r2,  "Higher is better (max=1)", "{:.3f}"),
]:
    colors = [MODEL_COLORS[m] for m in MODEL_ORDER]
    bars = ax.bar(MODEL_ORDER, vals.values, color=colors, edgecolor="white", linewidth=0.8, width=0.6)
    for bar, val in zip(bars, vals.values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + abs(vals.values.min())*0.02,
                fmt.format(val), ha="center", fontsize=8.5, color="#333")
    ax.set_ylabel(metric, **FONT_AXIS)
    ax.set_title(f"{metric} by Model — {label}", **FONT_TITLE)
    ax.tick_params(axis="x", rotation=20)
    if metric.startswith("Avg R"):
        ax.axhline(0, color="#999", linewidth=0.8, linestyle="--")

plt.tight_layout()
p2 = os.path.join(FIG_DIR, "02_model_comparison_rmse_r2.png")
plt.savefig(p2, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p2}")

# ---------------------------------------------------------------------------
# Figure 3 – R² Heatmap (Stocks × Models)
# ---------------------------------------------------------------------------
print("[3/8] R² heatmap …")
data_r2 = pv_r2[MODEL_ORDER].copy()
# Clip for color clarity: R² in [-3, 1]
data_clipped = data_r2.clip(-3, 1)
symbols_sorted = data_r2["CNN-LSTM"].sort_values(ascending=False).index

cmap_r2 = LinearSegmentedColormap.from_list("r2cmap", ["#d73027", "#fee08b", "#1a9850"])

fig, ax = plt.subplots(figsize=(12, max(8, len(symbols_sorted) * 0.3)))
im = ax.imshow(data_clipped.loc[symbols_sorted].T.values,
               aspect="auto", cmap=cmap_r2, vmin=-3, vmax=1)

ax.set_xticks(range(len(symbols_sorted)))
ax.set_xticklabels(symbols_sorted, rotation=90, fontsize=7)
ax.set_yticks(range(len(MODEL_ORDER)))
ax.set_yticklabels(MODEL_ORDER, fontsize=9)
ax.set_title("R² Score Heatmap — All Nifty 50 Stocks × All Models\n(Green = Good fit, Red = Poor fit)", **FONT_TITLE)

# Annotate cells
for i, model in enumerate(MODEL_ORDER):
    for j, sym in enumerate(symbols_sorted):
        val = data_r2.loc[sym, model] if sym in data_r2.index else np.nan
        if not np.isnan(val):
            txt = f"{val:.2f}"
            color = "white" if val < -1 else "black"
            ax.text(j, i, txt, ha="center", va="center", fontsize=5.5, color=color)

plt.colorbar(im, ax=ax, label="R² Score", fraction=0.02, pad=0.02)
plt.tight_layout()
p3 = os.path.join(FIG_DIR, "03_r2_heatmap.png")
plt.savefig(p3, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p3}")

# ---------------------------------------------------------------------------
# Figure 4 – MAPE Heatmap (Stocks × Models)
# ---------------------------------------------------------------------------
print("[4/8] MAPE heatmap …")
data_mape = pv_mape[MODEL_ORDER].copy()
symbols_mape = data_mape["CNN-LSTM"].sort_values().index   # sort by CNN-LSTM MAPE ascending
cmap_mape = LinearSegmentedColormap.from_list("mapecmap", ["#1a9850", "#fee08b", "#d73027"])

fig, ax = plt.subplots(figsize=(12, max(8, len(symbols_mape) * 0.3)))
im = ax.imshow(data_mape.loc[symbols_mape].T.values,
               aspect="auto", cmap=cmap_mape, vmin=0, vmax=30)

ax.set_xticks(range(len(symbols_mape)))
ax.set_xticklabels(symbols_mape, rotation=90, fontsize=7)
ax.set_yticks(range(len(MODEL_ORDER)))
ax.set_yticklabels(MODEL_ORDER, fontsize=9)
ax.set_title("MAPE (%) Heatmap — All Nifty 50 Stocks × All Models\n(Green = Low error, Red = High error)", **FONT_TITLE)

for i, model in enumerate(MODEL_ORDER):
    for j, sym in enumerate(symbols_mape):
        val = data_mape.loc[sym, model] if sym in data_mape.index else np.nan
        if not np.isnan(val):
            txt = f"{val:.1f}"
            color = "white" if val > 20 else "black"
            ax.text(j, i, txt, ha="center", va="center", fontsize=5.5, color=color)

plt.colorbar(im, ax=ax, label="MAPE (%)", fraction=0.02, pad=0.02)
plt.tight_layout()
p4 = os.path.join(FIG_DIR, "04_mape_heatmap.png")
plt.savefig(p4, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p4}")

# ---------------------------------------------------------------------------
# Figure 5 – Forecast Accuracy Distribution
# ---------------------------------------------------------------------------
print("[5/8] Forecast accuracy …")
fig, ax = plt.subplots(figsize=(10, 5))
for i, model in enumerate(MODEL_ORDER):
    vals = pv_acc[model].dropna().values
    bp = ax.boxplot(vals, positions=[i], widths=0.5,
                    patch_artist=True,
                    boxprops=dict(facecolor=MODEL_COLORS[model], alpha=0.75),
                    medianprops=dict(color="#1a1a2e", linewidth=2),
                    whiskerprops=dict(color="#777"),
                    capprops=dict(color="#777"),
                    flierprops=dict(marker="o", markersize=4, color="#999"))

# Remove the random baseline line as it doesn't make sense for generic accuracy
ax.set_xticks(range(len(MODEL_ORDER)))
ax.set_xticklabels(MODEL_ORDER, fontsize=10)
ax.set_ylabel("Forecast Accuracy (%)", **FONT_AXIS)
ax.set_title("Forecast Accuracy Distribution by Model — Nifty 50\n(Accuracy = 100 - MAPE | Box = IQR, Line = Median)", **FONT_TITLE)
ax.set_ylim(60, 100) # Accuracy is usually high
plt.tight_layout()
p5 = os.path.join(FIG_DIR, "05_forecast_accuracy.png")
plt.savefig(p5, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p5}")

# ---------------------------------------------------------------------------
# Figure 6 – F1-Score Heatmap
# ---------------------------------------------------------------------------
print("[6/8] F1-Score heatmap …")
data_f1 = pv_f1[MODEL_ORDER].copy()
symbols_f1 = data_f1["CNN-LSTM"].sort_values(ascending=False).index
cmap_f1 = LinearSegmentedColormap.from_list("f1cmap", ["#f7fbff", "#4393c3", "#053061"])

fig, ax = plt.subplots(figsize=(12, max(8, len(symbols_f1) * 0.3)))
im = ax.imshow(data_f1.loc[symbols_f1].T.values,
               aspect="auto", cmap=cmap_f1, vmin=0, vmax=0.7)

ax.set_xticks(range(len(symbols_f1)))
ax.set_xticklabels(symbols_f1, rotation=90, fontsize=7)
ax.set_yticks(range(len(MODEL_ORDER)))
ax.set_yticklabels(MODEL_ORDER, fontsize=9)
ax.set_title("F1-Score Heatmap — Direction Prediction (BUY/SELL)\nAll Nifty 50 Stocks × All Models", **FONT_TITLE)

for i, model in enumerate(MODEL_ORDER):
    for j, sym in enumerate(symbols_f1):
        val = data_f1.loc[sym, model] if sym in data_f1.index else np.nan
        if not np.isnan(val):
            ax.text(j, i, f"{val:.2f}", ha="center", va="center",
                    fontsize=5.5, color="white" if val > 0.4 else "black")

plt.colorbar(im, ax=ax, label="F1-Score", fraction=0.02, pad=0.02)
plt.tight_layout()
p6 = os.path.join(FIG_DIR, "06_f1_score_heatmap.png")
plt.savefig(p6, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p6}")

# ---------------------------------------------------------------------------
# Figure 7 – Top 15 stocks by CNN-LSTM R² (horizontal bar)
# ---------------------------------------------------------------------------
print("[7/8] Top 15 CNN-LSTM R² …")
cnn_r2 = cnn["R2"].dropna().sort_values(ascending=False).head(15)

fig, ax = plt.subplots(figsize=(9, 6))
colors_bar = ["#2E8B57" if v >= 0.9 else "#5CB878" if v >= 0.7 else "#A8D5B5" for v in cnn_r2.values]
bars = ax.barh(cnn_r2.index, cnn_r2.values, color=colors_bar, edgecolor="white")
for bar, val in zip(bars, cnn_r2.values):
    ax.text(val - 0.01, bar.get_y() + bar.get_height()/2,
            f"{val:.4f}", va="center", ha="right", fontsize=9,
            color="white" if val > 0.5 else "#333")

ax.axvline(0.9, color="#aaa", linewidth=1, linestyle="--", label="R²=0.90 threshold")
ax.set_xlabel("R² Score (CNN-LSTM)", **FONT_AXIS)
ax.set_title("Top 15 Nifty 50 Stocks — CNN-LSTM R² Score\n(Train: 13 yrs | Test: 2 yrs | Higher = Better)", **FONT_TITLE)
ax.invert_yaxis()
ax.set_xlim(0, 1.05)
ax.legend(fontsize=9)
patches = [
    mpatches.Patch(color="#2E8B57", label="R² ≥ 0.90 (Excellent)"),
    mpatches.Patch(color="#5CB878", label="R² ≥ 0.70 (Good)"),
    mpatches.Patch(color="#A8D5B5", label="R² < 0.70 (Moderate)"),
]
ax.legend(handles=patches, fontsize=8, loc="lower right")
plt.tight_layout()
p7 = os.path.join(FIG_DIR, "07_best_model_r2_top15.png")
plt.savefig(p7, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p7}")

# ---------------------------------------------------------------------------
# Figure 8 – CNN-LSTM vs LSTM R² scatter
# ---------------------------------------------------------------------------
print("[8/8] CNN-LSTM vs LSTM R² scatter …")
lstm_r2  = df_raw[df_raw["Model"] == "LSTM"].set_index("Symbol")["R2"]
cnn_r2_s = df_raw[df_raw["Model"] == "CNN-LSTM"].set_index("Symbol")["R2"]
common = lstm_r2.index.intersection(cnn_r2_s.index)

fig, ax = plt.subplots(figsize=(8, 7))
x = lstm_r2.loc[common].values
y = cnn_r2_s.loc[common].values
ax.scatter(x, y, s=60, alpha=0.75, c="#2E8B57", edgecolors="white", linewidth=0.8)
for sym in common:
    if abs(cnn_r2_s[sym] - lstm_r2[sym]) > 0.15 or cnn_r2_s[sym] > 0.85:
        ax.annotate(sym, (lstm_r2[sym], cnn_r2_s[sym]),
                    textcoords="offset points", xytext=(5, 4), fontsize=7, color="#555")

lims = [min(x.min(), y.min()) - 0.1, max(x.max(), y.max()) + 0.1]
ax.plot(lims, lims, "k--", linewidth=1, alpha=0.4, label="LSTM = CNN-LSTM line")
ax.axhline(0, color="#ddd", linewidth=0.8); ax.axvline(0, color="#ddd", linewidth=0.8)
ax.set_xlabel("LSTM R² Score", **FONT_AXIS)
ax.set_ylabel("CNN-LSTM R² Score", **FONT_AXIS)
ax.set_title("CNN-LSTM vs LSTM — R² Comparison\n(Points above diagonal = CNN-LSTM wins)", **FONT_TITLE)
ax.legend(fontsize=9)
above = np.sum(y > x)
ax.text(0.05, 0.95, f"CNN-LSTM wins: {above}/{len(common)} stocks",
        transform=ax.transAxes, fontsize=10, color="#2E8B57",
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#2E8B57", alpha=0.8))
plt.tight_layout()
p8 = os.path.join(FIG_DIR, "08_cnnlstm_vs_lstm_scatter.png")
plt.savefig(p8, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Saved: {p8}")

# ---------------------------------------------------------------------------
# Excel workbook with formatted tables
# ---------------------------------------------------------------------------
print("\nGenerating Excel workbook …")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # --- Sheet 1: Master data ---
    ws1 = wb.active
    ws1.title = "All Models All Stocks"
    header_fill = PatternFill("solid", fgColor="1A3C5A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(bottom=Side(style="thin", color="CCCCCC"))

    model_fills = {
        "ARIMA":         "A8C4E0",
        "SARIMA":        "6FA8D0",
        "ARIMA-LSTM": "F4D0A0",
        "XGBoost":       "F4A460",
        "LSTM":          "B0D8B0",
        "CNN-LSTM":      "2E8B57",
    }
    model_fonts = {"CNN-LSTM": Font(color="FFFFFF", bold=True, size=9)}

    cols = ["Symbol", "Model", "RMSE", "MAE", "MAPE", "Accuracy (%)", "R2",
            "Directional_Accuracy", "Precision", "Recall", "F1_Score"]
    df_export = df_raw[cols].copy()
    df_export = df_export.sort_values(["Symbol", "Model"])

    for ci, col in enumerate(cols, 1):
        cell = ws1.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(df_export.itertuples(index=False), 2):
        model = row.Model
        fill  = PatternFill("solid", fgColor=model_fills.get(model, "FFFFFF"))
        font  = model_fonts.get(model, Font(size=9))
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci)
            if isinstance(val, float):
                cell.value  = round(val, 4)
                cell.number_format = "0.0000"
            else:
                cell.value = val
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border

    # Auto column widths
    for ci in range(1, len(cols) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    # --- Sheet 2: CNN-LSTM Summary ---
    ws2 = wb.create_sheet("CNN-LSTM Summary")
    sum_cols = ["Symbol", "RMSE", "MAE", "MAPE", "Accuracy (%)", "R2",
                "Directional_Accuracy", "Precision", "Recall", "F1_Score"]
    cnn_df = df_raw[df_raw["Model"] == "CNN-LSTM"][sum_cols].sort_values("R2", ascending=False)

    for ci, col in enumerate(sum_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = PatternFill("solid", fgColor="2E8B57")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(cnn_df.itertuples(index=False), 2):
        r2val = row.R2
        if r2val > 0.9:
            row_fill = PatternFill("solid", fgColor="C8EFCE")
        elif r2val > 0.7:
            row_fill = PatternFill("solid", fgColor="E8F5E9")
        elif r2val > 0:
            row_fill = PatternFill("solid", fgColor="FFF8E1")
        else:
            row_fill = PatternFill("solid", fgColor="FFEBEE")

        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci)
            if isinstance(val, float):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
            else:
                cell.value = val
            cell.fill      = row_fill
            cell.font      = Font(size=9)
            cell.alignment = Alignment(horizontal="center")

    for ci in range(1, len(sum_cols) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 17

    # --- Sheet 3: Model average summary ---
    ws3 = wb.create_sheet("Model Average Summary")
    avg_df = df_raw.groupby("Model")[["RMSE","MAE","MAPE","Accuracy (%)","R2",
                                      "Directional_Accuracy","Precision","Recall","F1_Score"]].mean()
    avg_df = avg_df.reindex(MODEL_ORDER).reset_index()

    for ci, col in enumerate(avg_df.columns, 1):
        cell = ws3.cell(row=1, column=ci, value=col)
        cell.fill = PatternFill("solid", fgColor="1A3C5A")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(avg_df.itertuples(index=False), 2):
        model = row.Model
        fill  = PatternFill("solid", fgColor=model_fills.get(model, "FFFFFF").lstrip("#") if model_fills.get(model,"#FFFFFF").startswith("#") else model_fills.get(model,"FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=ci)
            if isinstance(val, float):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
            else:
                cell.value = val
            cell.font      = Font(size=9, bold=(model=="CNN-LSTM"))
            cell.alignment = Alignment(horizontal="center")

    for ci in range(1, len(avg_df.columns) + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 18

    wb.save(EXCEL_PATH)
    print(f"  Saved: {EXCEL_PATH}")
    excel_ok = True
except ImportError:
    print("  openpyxl not installed — skipping Excel export.")
    excel_ok = False

# ---------------------------------------------------------------------------
# HTML Report — filled from report_template.html
# ---------------------------------------------------------------------------
print("\nGenerating HTML report …")

TEMPLATE_PATH = "report_template.html"

def fig_tag(path, caption, width="100%"):
    """Return an HTML <figure> tag with a relative img path."""
    rel = os.path.relpath(path, REPORT_DIR).replace("\\", "/")
    return (
        f'\n    <figure>\n'
        f'      <img src="{rel}" alt="{caption}" '
        f'style="width:{width};border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.12);">\n'
        f'      <figcaption>{caption}</figcaption>\n'
        f'    </figure>'
    )

def df_to_html(df, highlight_col=None, green_if_high=True):
    """Convert a DataFrame to a styled HTML table string."""
    rows = ""
    for _, row in df.iterrows():
        cells = ""
        for col in df.columns:
            val   = row[col]
            style = ""
            if highlight_col and col == highlight_col and isinstance(val, float):
                if green_if_high:
                    bg = ("#c8efce" if val > 0.9 else
                          "#e8f5e9" if val > 0.7 else
                          "#fff8e1" if val > 0   else "#ffebee")
                else:
                    bg = ("#c8efce" if val < 5  else
                          "#e8f5e9" if val < 10  else
                          "#fff8e1" if val < 20  else "#ffebee")
                style = f'style="background:{bg}"'
            display = f"{val:.4f}" if isinstance(val, float) else str(val)
            cells  += f"<td {style}>{display}</td>"
        rows += f"<tr>{cells}</tr>\n"

    headers = "".join(f"<th>{c}</th>" for c in df.columns)
    return f"<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>"

# Build data for the template
avg_df_display = (
    df_raw.groupby("Model")[
        ["RMSE", "MAE", "MAPE", "Accuracy (%)", "R2",
         "Directional_Accuracy", "Precision", "Recall", "F1_Score"]
    ].mean()
    .reindex(MODEL_ORDER)
    .round(4)
    .reset_index()
)

cnn_top = (
    df_raw[df_raw["Model"] == "CNN-LSTM"][[
        "Symbol", "RMSE", "MAE", "MAPE", "Accuracy (%)", "R2",
        "Directional_Accuracy", "Precision", "Recall", "F1_Score"
    ]]
    .sort_values("R2", ascending=False)
    .round(4)
)

# Placeholder → value mapping
cnn = df_raw[df_raw["Model"] == "CNN-LSTM"]
replacements = {
    "{{NUM_SYMBOLS}}":    str(df_raw["Symbol"].nunique()),
    "{{CNN_AVG_R2}}":     f"{cnn['R2'].mean():.3f}",
    "{{CNN_AVG_MAPE}}":   f"{cnn['MAPE'].mean():.2f}",
    "{{R2_ABOVE_90}}":    str(int((cnn["R2"] > 0.9).sum())),
    "{{R2_ABOVE_0}}":     str(int((cnn["R2"] > 0.0).sum())),
    "{{TABLE_AVERAGES}}": df_to_html(avg_df_display),
    "{{TABLE_CNN_LSTM}}": df_to_html(cnn_top.reset_index(drop=True),
                                     highlight_col="R2", green_if_high=True),
    "{{FIG_1}}": fig_tag(p1, "Figure 1: Average MAPE by Model — lower is better"),
    "{{FIG_2}}": fig_tag(p2, "Figure 2: Average RMSE and R² by Model"),
    "{{FIG_3}}": fig_tag(p3, "Figure 3: R² Score Heatmap (Green=Good, Red=Poor). "
                              "Sorted by CNN-LSTM R² descending.", "100%"),
    "{{FIG_4}}": fig_tag(p4, "Figure 4: MAPE % Heatmap (Green=Low error, Red=High error). "
                              "Sorted by CNN-LSTM MAPE ascending.", "100%"),
    "{{FIG_5}}": fig_tag(p5, "Figure 5: Distribution of Forecast Accuracy (100 - MAPE) by Model "
                              "across all 50 stocks."),
    "{{FIG_6}}": fig_tag(p6, "Figure 6: F1-Score Heatmap for BUY/SELL classification. "
                              "Darker blue = better balance between precision and recall.", "100%"),
    "{{FIG_7}}": fig_tag(p7, "Figure 7: Top 15 Nifty 50 stocks where CNN-LSTM achieves "
                              "best R² (train:13yr / test:2yr)"),
    "{{FIG_8}}": fig_tag(p8, "Figure 8: CNN-LSTM vs LSTM R² scatter — "
                              "points above diagonal = CNN-LSTM wins"),
}

# Load template and substitute
if not os.path.exists(TEMPLATE_PATH):
    print(f"  ERROR: Template not found at '{TEMPLATE_PATH}'")
    print("  Make sure report_template.html is in the same folder as generate_report.py")
else:
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        html_out = f.read()

    for placeholder, value in replacements.items():
        html_out = html_out.replace(placeholder, value)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html_out)

    print(f"  Template : {TEMPLATE_PATH}")
    print(f"  Saved    : {HTML_PATH}")

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print()
print("=" * 60)
print("  REPORT GENERATION COMPLETE")
print("=" * 60)
print(f"  Template : {TEMPLATE_PATH}")
print(f"  Figures  : {FIG_DIR}/  (8 PNG charts)")
if excel_ok:
    print(f"  Excel    : {EXCEL_PATH}")
print(f"  HTML     : {HTML_PATH}")
print()
print("  Open dissertation_report.html in your browser to view")
print("  the full interactive report with all charts and tables.")
print("=" * 60)

